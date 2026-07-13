"""
Budget {Y} workbook builder.

Produces an .xlsx mirroring the layout of the manually-maintained
"Budget {Y} avec légende" Google Sheet (see ``Budget 2026.xlsx`` reference):
a single block (rows 11-25) with Légende, the ``Au DD/MM/YYYY`` date, headers
per BU, then Devis/Contrats Signés/Potentiels/Envoyés numbers, Production
sécurisée, Portefeuille sites au {today}, Total sécurisé maintenance,
Objectifs and Production à aller chercher.

All numeric inputs (BU sums, portefeuille values) are computed from the
processed proposals DataFrames the dashboard already loads — no new business
logic, just aggregation + layout.
"""

from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any, Dict, List, Optional, Sequence

import pandas as pd

try:
    from config.settings import STATUS_WAITING, STATUS_WON
except ImportError:
    STATUS_WON = ['gagné', 'gagne', 'signé', 'signe', 'gagnés et finis', 'gagnés en cours']
    STATUS_WAITING = ['brief', 'en cours', 'envoyée(s) attente réponse', 'envoyée(s) en attente de réponse']

try:
    from src.processing.objectives import objective_for_year
except ImportError:
    def objective_for_year(year: int, metric: str, dimension: str, key: str) -> float:  # type: ignore
        return 0.0


BU_ORDER: List[str] = ['CONCEPTION', 'TRAVAUX', 'MAINTENANCE']

LEGEND_TEXT_TEMPLATE = (
    "Devis Signés = Tous les devis déjà signés à produire en {year}\n"
    "Devis Potentiels = Tous les devis envoyés multipliés par leur probabilité "
    "de gain à produire en {year}\n"
    "Devis Envoyés = Tous les devis envoyés sans probabilité à produire en {year}"
)

EUR_FORMAT = '#,##0 €'


# =============================================================================
# Aggregation helpers
# =============================================================================

def _normalize_status(df: pd.DataFrame) -> pd.Series:
    """Return a normalized lowercase status Series (statut_clean fallback to statut)."""
    if 'statut_clean' in df.columns:
        return df['statut_clean'].astype(str).str.strip().str.lower()
    if 'statut' in df.columns:
        return df['statut'].astype(str).str.strip().str.lower()
    return pd.Series([''] * len(df), index=df.index)


def _bu_column(df: pd.DataFrame) -> pd.Series:
    """Return BU Series (final_bu fallback to cf_bu)."""
    if 'final_bu' in df.columns:
        return df['final_bu'].astype(str).str.strip().str.upper()
    if 'cf_bu' in df.columns:
        return df['cf_bu'].astype(str).str.strip().str.upper()
    return pd.Series([''] * len(df), index=df.index)


def _to_year_set(values: Sequence[str]) -> set:
    return {v.strip().lower() for v in values}


def _safe_sum(series: pd.Series) -> float:
    if series is None or len(series) == 0:
        return 0.0
    return float(pd.to_numeric(series, errors='coerce').fillna(0.0).sum())


def _column(df: pd.DataFrame, name: str) -> pd.Series:
    """Return df[name] coerced to numeric, or zeros if column missing."""
    if name in df.columns:
        return pd.to_numeric(df[name], errors='coerce').fillna(0.0)
    return pd.Series([0.0] * len(df), index=df.index)


def _compute_bu_amounts(df: pd.DataFrame, year: int, bu: str) -> Dict[str, float]:
    """
    Compute Signés / Potentiels / Envoyés totals for one BU and one year.

    Definitions (won and sent pipes stay disjoint — they never mix):
    - Signés     = sum(Montant Total {Y})   where status ∈ STATUS_WON     and BU == bu
    - Potentiels = sum(Montant Pondéré {Y}) where status ∈ STATUS_WAITING and BU == bu
    - Envoyés    = sum(Montant Total {Y})   where status ∈ STATUS_WAITING and BU == bu
    """
    if df is None or df.empty:
        return {"signes": 0.0, "potentiels": 0.0, "envoyes": 0.0}

    won = _to_year_set(STATUS_WON)
    waiting = _to_year_set(STATUS_WAITING)
    statuses = _normalize_status(df)
    bus = _bu_column(df)

    bu_mask = bus == bu.upper()
    won_mask = bu_mask & statuses.isin(won)
    waiting_mask = bu_mask & statuses.isin(waiting)

    total_col = f"Montant Total {year}"
    pondere_col = f"Montant Pondéré {year}"

    signes = _safe_sum(_column(df, total_col)[won_mask])
    potentiels = _safe_sum(_column(df, pondere_col)[waiting_mask])
    envoyes = _safe_sum(_column(df, total_col)[waiting_mask])

    return {"signes": signes, "potentiels": potentiels, "envoyes": envoyes}


def _sum_production_by_bu(df: pd.DataFrame, year: int, bu: str, weighted: bool = False) -> float:
    """
    Sum a production column (`Montant Total {Y}` or `Montant Pondéré {Y}`) for one BU.

    Used with production-aggregated data (``load_aggregated_production_data``), which
    already pulls the Signé/Envoyé sheets across signing years Y-2..Y and keeps only
    rows producing in {Y}. This is how prior-year signatures cascade into year {Y}.
    No status filter is needed: Signé sheets are all won, Envoyé sheets all waiting.
    """
    if df is None or df.empty:
        return 0.0
    col = f"Montant {'Pondéré' if weighted else 'Total'} {year}"
    bus = _bu_column(df)
    return _safe_sum(_column(df, col)[bus == bu.upper()])


def dedupe_sent_pipe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Drop duplicate devis from the aggregated SENT pipe, keeping the freshest copy.

    Since the Envoyé sheets became a live snapshot (keyed on devis date), a pending
    devis dated {Y} whose Furious record was created in {Y-1} appears twice in the
    aggregation: once in the frozen {Y-1} monthly sheets and once in the live {Y}
    sheets. Keep the copy from the highest ``signed_year`` (the live one — fresher
    status and production columns); without the column, keep the last occurrence.
    """
    if df is None or df.empty or 'id' not in df.columns:
        return df
    if 'signed_year' in df.columns:
        order = pd.to_numeric(df['signed_year'], errors='coerce').fillna(0)
        df = df.iloc[order.argsort(kind='stable')]
    return df.drop_duplicates(subset=['id'], keep='last')


def filter_carryover_by_pending(
    df: pd.DataFrame,
    budget_year: int,
    pending_ids: Optional[set],
) -> pd.DataFrame:
    """
    Keep prior-year carryover rows only if the devis is still pending in Furious.

    Rows from the frozen {Y-1} sheets have their status frozen at write time — a
    devis lost months ago still reads "en attente" there. ``pending_ids`` is the
    set of currently-WAITING devis ids written daily by the reconciliation sidecar
    (data/pending_ids.json). Current-year rows are untouched: the live sheets are
    rewritten daily so their statuses are already fresh.

    With ``pending_ids`` None (store missing/stale), returns df unchanged so the
    budget still generates — the projet_start pruning remains as a safety net.
    """
    if df is None or df.empty or pending_ids is None:
        return df
    if 'signed_year' not in df.columns or 'id' not in df.columns:
        return df
    signed_year = pd.to_numeric(df['signed_year'], errors='coerce')
    is_carryover = signed_year < budget_year
    still_pending = df['id'].astype(str).str.strip().isin({str(p) for p in pending_ids})
    return df[~is_carryover | still_pending].copy()


def drop_stale_sent_carryover(df: pd.DataFrame, budget_year: int, today: date) -> pd.DataFrame:
    """
    Remove stale carried-over proposals from the SENT (Envoyé) pipe.

    Drops rows that were sent in a previous year (``signed_year`` < budget_year)
    *and* whose ``projet_start`` is already overdue (before ``today``): a project
    that should have started but is still unsigned is no longer a realistic
    contributor to {budget_year} production.

    Current-year-sent proposals are always kept, as are prior-year proposals
    whose start is still in the future (or unknown).
    """
    if df is None or df.empty:
        return df
    if 'signed_year' not in df.columns or 'projet_start' not in df.columns:
        return df

    signed_year = pd.to_numeric(df['signed_year'], errors='coerce')
    start = pd.to_datetime(df['projet_start'], errors='coerce')
    today_ts = pd.Timestamp(today)

    is_carryover = signed_year < budget_year
    is_overdue = start.notna() & (start < today_ts)
    stale = is_carryover & is_overdue
    return df[~stale].copy()


# =============================================================================
# Workbook layout
# =============================================================================

def _set_eur(cell, value: Optional[float]) -> None:
    if value is None:
        cell.value = ""
    else:
        cell.value = float(value)
        cell.number_format = EUR_FORMAT


def _build_sheet1(ws, year: int, today: date, bu_totals: Dict[str, Dict[str, float]],
                  portefeuille_debut: Optional[float], portefeuille_running: Optional[float]) -> None:
    """
    Layout (1-indexed rows) — mirrors `Copie de Budget 26 avec légende`:

      D11           : "Légende"
      D12:L14       : legend text (merged)
      D16           : "Au DD/MM/YYYY"
      D17:D19       : "Projection {Y}" (merged)
      E17:G17       : "CONCEPTION"  (merged)
      H17:J17       : "TRAVAUX"     (merged)
      K17:M17       : "MAINTENANCE" (merged)
      N17:P17       : "TOTAL"       (merged)
      Row 18        : column subheaders
      Row 19        : computed numeric values
      D20           : "Production sécurisée"
      Row 20        : E20=E19+F19, H20=H19+I19, K20=L19+K19, N20=E20+H20+L22
      K21:M21       : "Portefeuille sites au DD/MM/YYYY"
      L21           : portefeuille_running
      K22:M22       : "Total sécurisé maintenance {YY}"
      L22           : =L19+L21
      D23           : "Objectifs {Y}"
      Row 23        : E23/H23/K23 from objective_for_year, N23 = sum
      D25           : "Production à aller chercher"
      Row 25        : =Obj-Production sécurisée per BU
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    ws.title = f"Budget {year} avec légende"

    # Column widths matching the reference workbook
    width_map = {
        'D': 28.1, 'E': 13.9, 'F': 16.0, 'G': 17.0, 'H': 14.0, 'I': 16.0, 'J': 17.0,
        'K': 18.0, 'L': 16.4, 'M': 14.9, 'N': 21.5, 'O': 23.1, 'P': 22.2,
    }
    for col, w in width_map.items():
        ws.column_dimensions[col].width = w

    bold = Font(bold=True)
    bold_white = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    legend_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # BU band colors aligned with the dashboard BU_COLORS (source of truth):
    # CONCEPTION green, TRAVAUX gold, MAINTENANCE purple; TOTAL blue.
    bu_fills = {
        'CONCEPTION': PatternFill("solid", fgColor="2D5A3F"),
        'TRAVAUX':    PatternFill("solid", fgColor="F4C430"),
        'MAINTENANCE': PatternFill("solid", fgColor="7B4B94"),
        'TOTAL':      PatternFill("solid", fgColor="3D85C6"),
    }
    # Light tints for the sub-header row (row 18), matching the original workbook.
    bu_sub_fills = {
        'CONCEPTION': PatternFill("solid", fgColor="D9EAD3"),
        'TRAVAUX':    PatternFill("solid", fgColor="FFF2CC"),
        'MAINTENANCE': PatternFill("solid", fgColor="D9D2E9"),
        'TOTAL':      PatternFill("solid", fgColor="CFE2F3"),
    }
    # Dark bands use white text; the gold TRAVAUX band stays black for contrast.
    bu_fonts = {
        'CONCEPTION': bold_white,
        'TRAVAUX':    bold,
        'MAINTENANCE': bold_white,
        'TOTAL':      bold_white,
    }

    # --- Légende block --------------------------------------------------
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont

    ws['D11'] = 'Légende'
    ws['D11'].font = bold
    ws.merge_cells('E11:L11')   # blank header bar next to "Légende"
    ws.merge_cells('D12:L14')
    bold_term = InlineFont(b=True)
    ws['D12'] = CellRichText(
        TextBlock(bold_term, "Devis Signés"),
        f" = Tous les devis déjà signés à produire en {year}\n",
        TextBlock(bold_term, "Devis Potentiels"),
        f" = Tous les devis envoyés multipliés par leur probabilité de gain à produire en {year}\n",
        TextBlock(bold_term, "Devis Envoyés"),
        f" = Tous les devis envoyés sans probabilité à produire en {year}",
    )
    ws['D12'].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    # Outline + fill the whole légende box (title row + body), like the table.
    for row in range(11, 15):
        for col_idx in range(4, 13):  # D..L
            cell = ws.cell(row=row, column=col_idx)
            cell.border = border
            if row >= 12:
                cell.fill = legend_fill

    # --- Date row -------------------------------------------------------
    ws['D16'] = f"Au {today.strftime('%d/%m/%Y')}"
    ws['D16'].font = bold

    # --- Header band ----------------------------------------------------
    ws.merge_cells('D17:D19')
    ws['D17'] = f"Projection {year}"
    ws['D17'].font = bold
    ws['D17'].alignment = center

    bu_col_groups = {
        'CONCEPTION': ('E', 'F', 'G'),
        'TRAVAUX': ('H', 'I', 'J'),
        'MAINTENANCE': ('K', 'L', 'M'),
        'TOTAL': ('N', 'O', 'P'),
    }
    for label, cols in bu_col_groups.items():
        rng = f"{cols[0]}17:{cols[2]}17"
        ws.merge_cells(rng)
        first = f"{cols[0]}17"
        ws[first] = label
        ws[first].font = bu_fonts[label]
        ws[first].alignment = center
        ws[first].fill = bu_fills[label]

    # --- Subheaders (row 18) -------------------------------------------
    sub_headers = {
        'E18': 'Devis Signés',     'F18': 'Devis Potentiels', 'G18': 'Devis Envoyés',
        'H18': 'Devis Signés',     'I18': 'Devis Potentiels', 'J18': 'Devis Envoyés',
        'K18': f'Nouveaux contrats {year}', 'L18': 'Contrats Potentiels', 'M18': 'Contrats Envoyés',
        'N18': 'Devis + Contrats Signés', 'O18': 'Devis + Contrats Potentiels', 'P18': 'Devis + Contrats Envoyés',
    }
    for coord, label in sub_headers.items():
        ws[coord] = label
        ws[coord].font = bold
        ws[coord].alignment = center
    for label, cols in bu_col_groups.items():
        for c in cols:
            ws[f"{c}18"].fill = bu_sub_fills[label]

    # --- Row 19 numeric values -----------------------------------------
    cells_19 = [
        ('E19', bu_totals['CONCEPTION']['signes']),
        ('F19', bu_totals['CONCEPTION']['potentiels']),
        ('G19', bu_totals['CONCEPTION']['envoyes']),
        ('H19', bu_totals['TRAVAUX']['signes']),
        ('I19', bu_totals['TRAVAUX']['potentiels']),
        ('J19', bu_totals['TRAVAUX']['envoyes']),
        ('K19', bu_totals['MAINTENANCE']['signes']),
        ('L19', bu_totals['MAINTENANCE']['potentiels']),
        ('M19', bu_totals['MAINTENANCE']['envoyes']),
    ]
    for coord, value in cells_19:
        _set_eur(ws[coord], value)
    ws['N19'] = '=E19+H19+K19'
    ws['O19'] = '=F19+I19+L19'
    ws['P19'] = '=G19+J19+M19'
    for coord in ('N19', 'O19', 'P19'):
        ws[coord].number_format = EUR_FORMAT

    # --- Row 20 Production sécurisée -----------------------------------
    ws['D20'] = 'Production sécurisée'
    ws['D20'].font = bold
    ws['E20'] = '=E19+F19'
    ws['H20'] = '=H19+I19'
    ws['K20'] = '=L19+K19'
    ws['N20'] = '=E20+H20+L22'
    for coord in ('E20', 'H20', 'K20', 'N20'):
        ws[coord].number_format = EUR_FORMAT
        ws[coord].font = bold
    for rng in ('E20:G20', 'H20:J20', 'K20:M20', 'N20:P20'):
        ws.merge_cells(rng)

    maint_fill = PatternFill("solid", fgColor="D9D2E9")  # MAINTENANCE light purple

    wrap_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # --- Row 21 Portefeuille sites au {today} --------------------------
    ws['K21'] = f"Portefeuille sites au {today.strftime('%d/%m/%Y')}: "
    ws['K21'].alignment = wrap_right
    if portefeuille_running is not None:
        _set_eur(ws['L21'], portefeuille_running)
    ws.merge_cells('L21:M21')
    for coord in ('K21', 'L21', 'M21'):
        ws[coord].fill = maint_fill

    # --- Row 22 Total sécurisé maintenance {YY} ------------------------
    ws['K22'] = f"Total sécurisé maintenance {str(year)[-2:]}"
    ws['K22'].alignment = wrap_right
    ws['K22'].font = bold
    ws['L22'] = '=L19+L21'
    ws['L22'].number_format = EUR_FORMAT
    ws['L22'].font = bold
    ws.merge_cells('L22:M22')
    for coord in ('K22', 'L22', 'M22'):
        ws[coord].fill = maint_fill

    # Merge the blank blocks flanking the maintenance portefeuille rows so the
    # CONCEPTION/TRAVAUX and TOTAL sides read as single cells (like the original).
    ws.merge_cells('D21:J22')
    ws.merge_cells('N21:P22')

    # Give rows 21-22 room for the wrapped maintenance labels (2 lines).
    ws.row_dimensions[21].height = 30
    ws.row_dimensions[22].height = 30

    # --- Row 23 Objectifs {Y} ------------------------------------------
    ws['D23'] = f"Objectifs {year}"
    ws['D23'].font = bold

    obj_conception = objective_for_year(year, 'signe', 'bu', 'CONCEPTION')
    obj_travaux = objective_for_year(year, 'signe', 'bu', 'TRAVAUX')
    obj_maintenance = objective_for_year(year, 'signe', 'bu', 'MAINTENANCE')

    _set_eur(ws['E23'], obj_conception)
    _set_eur(ws['H23'], obj_travaux)
    _set_eur(ws['K23'], obj_maintenance)
    ws['N23'] = '=E23+H23+K23'
    ws['N23'].number_format = EUR_FORMAT
    for rng in ('E23:G23', 'H23:J23', 'K23:M23', 'N23:P23'):
        ws.merge_cells(rng)

    # Blank separator row 24 as one block.
    ws.merge_cells('D24:P24')

    # --- Row 25 Production à aller chercher ----------------------------
    ws['D25'] = 'Production à aller chercher'
    ws['D25'].font = bold
    ws['E25'] = '=E23-E20'
    ws['H25'] = '=H23-H20'
    ws['K25'] = '=K23-L22'
    ws['N25'] = '=E25+H25+K25'
    for coord in ('E25', 'H25', 'K25', 'N25'):
        ws[coord].number_format = EUR_FORMAT
        ws[coord].font = bold
    for rng in ('E25:G25', 'H25:J25', 'K25:M25', 'N25:P25'):
        ws.merge_cells(rng)

    # --- Borders on the whole projection grid (D17:P25) ----------------
    for row in range(17, 26):
        for col_idx in range(4, 17):  # D..P
            ws.cell(row=row, column=col_idx).border = border


# =============================================================================
# Public entry point
# =============================================================================

def build_budget_workbook(
    year: int,
    df_processed: Optional[pd.DataFrame] = None,
    portefeuille_debut_annee: Optional[float] = None,
    portefeuille_running: Optional[float] = None,
    today: Optional[date] = None,
    *,
    bu_totals: Optional[Dict[str, Dict[str, float]]] = None,
) -> bytes:
    """
    Build the Budget {year} xlsx and return its bytes.

    Args:
        year: Target budget year (e.g. 2026).
        df_processed: Single-DataFrame legacy path (status-filtered) used only
            when ``bu_totals`` is not provided.
        portefeuille_debut_annee: Maintenance Entretien début {year}
            (sum of "Total HT Cette année" in Notion, start-of-year snapshot).
        portefeuille_running: Running maintenance portefeuille at click time
            (live Notion sum) — shown as "Portefeuille sites au {today}".
        today: Date stamped in the workbook headers (defaults to ``date.today()``).
        bu_totals: Precomputed {BU: {signes, potentiels, envoyes}} (preferred —
            lets the caller use production-year aggregation with carryover).
    """
    from openpyxl import Workbook

    today = today or date.today()

    if bu_totals is None:
        bu_totals = {bu: _compute_bu_amounts(df_processed, year, bu) for bu in BU_ORDER}

    wb = Workbook()
    ws1 = wb.active
    _build_sheet1(ws1, year, today, bu_totals, portefeuille_debut_annee, portefeuille_running)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

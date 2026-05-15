"""Tests for the Budget {Y} workbook builder."""

from datetime import date
from io import BytesIO

import pandas as pd
import pytest

from src.integrations.budget_export import (
    LEGEND_TEXT_TEMPLATE,
    _compute_bu_amounts,
    _compute_maintenance_entries,
    build_budget_workbook,
)


@pytest.fixture
def sample_df() -> pd.DataFrame:
    """
    Minimal processed-style DataFrame with one row per (BU, status) combination
    plus two extras to test signature-year filtering on the Maintenance sheet.

    Columns chosen to match what `read_worksheet` returns + what RevenueEngine adds.
    """
    return pd.DataFrame([
        # CONCEPTION won (Signé) — contributes to signes & envoyes
        {
            "id_devis": "C-W-1", "title": "Concept Won 1",
            "statut_clean": "gagné", "final_bu": "CONCEPTION", "amount": 100.0,
            "Montant Total 2026": 90.0, "Montant Pondéré 2026": 81.0,
            "signature_date": "2026-03-15", "date": "2026-03-15", "projet_start": "2026-04-01",
        },
        # CONCEPTION waiting (Envoyé) — contributes to potentiels & envoyes
        {
            "id_devis": "C-WT-1", "title": "Concept Waiting 1",
            "statut_clean": "en cours", "final_bu": "CONCEPTION", "amount": 200.0,
            "Montant Total 2026": 180.0, "Montant Pondéré 2026": 90.0,
            "signature_date": None, "date": "2026-02-01", "projet_start": "2026-05-01",
        },
        # TRAVAUX won
        {
            "id_devis": "T-W-1", "title": "Travaux Won 1",
            "statut_clean": "signé", "final_bu": "TRAVAUX", "amount": 1000.0,
            "Montant Total 2026": 950.0, "Montant Pondéré 2026": 950.0,
            "signature_date": "2026-01-10", "date": "2026-01-10", "projet_start": "2026-02-01",
        },
        # TRAVAUX waiting
        {
            "id_devis": "T-WT-1", "title": "Travaux Waiting 1",
            "statut_clean": "envoyée(s) en attente de réponse",
            "final_bu": "TRAVAUX", "amount": 500.0,
            "Montant Total 2026": 500.0, "Montant Pondéré 2026": 250.0,
            "signature_date": None, "date": "2026-04-01", "projet_start": "2026-05-01",
        },
        # MAINTENANCE won this year — entry on Maintenance sheet
        {
            "id_devis": "M-W-1", "title": "Maintenance Won 1",
            "statut_clean": "gagné", "final_bu": "MAINTENANCE", "amount": 7088.10,
            "Montant Total 2026": 7088 / 12 * 10, "Montant Pondéré 2026": 7088 / 12 * 10,
            "signature_date": "2026-03-20", "date": "2026-03-20", "projet_start": "2026-03-15",
        },
        # MAINTENANCE waiting
        {
            "id_devis": "M-WT-1", "title": "Maintenance Waiting 1",
            "statut_clean": "brief", "final_bu": "MAINTENANCE", "amount": 2000.0,
            "Montant Total 2026": 2000.0, "Montant Pondéré 2026": 1000.0,
            "signature_date": None, "date": "2026-06-01", "projet_start": "2026-07-01",
        },
        # MAINTENANCE won but in PREVIOUS year — must be excluded from maintenance entries
        {
            "id_devis": "M-W-OLD", "title": "Maintenance Won Last Year",
            "statut_clean": "gagné", "final_bu": "MAINTENANCE", "amount": 4444.0,
            "Montant Total 2026": 0.0, "Montant Pondéré 2026": 0.0,
            "signature_date": "2025-11-01", "date": "2025-11-01", "projet_start": "2025-12-01",
        },
    ])


# ---------------------------------------------------------------------------
# _compute_bu_amounts
# ---------------------------------------------------------------------------

def test_compute_bu_amounts_signes_uses_won_montant_total(sample_df):
    out = _compute_bu_amounts(sample_df, 2026, 'CONCEPTION')
    assert out['signes'] == pytest.approx(90.0)


def test_compute_bu_amounts_potentiels_uses_waiting_montant_pondere(sample_df):
    out = _compute_bu_amounts(sample_df, 2026, 'CONCEPTION')
    assert out['potentiels'] == pytest.approx(90.0)


def test_compute_bu_amounts_envoyes_uses_won_plus_waiting_montant_total(sample_df):
    out = _compute_bu_amounts(sample_df, 2026, 'CONCEPTION')
    # Won 90 + Waiting 180 (both Montant Total 2026)
    assert out['envoyes'] == pytest.approx(270.0)


def test_compute_bu_amounts_maintenance_excludes_other_bus(sample_df):
    out = _compute_bu_amounts(sample_df, 2026, 'MAINTENANCE')
    # Won this year: 7088/12*10 ; Won previous year: 0 in Montant Total 2026
    expected_signes = 7088 / 12 * 10
    assert out['signes'] == pytest.approx(expected_signes)
    assert out['potentiels'] == pytest.approx(1000.0)
    assert out['envoyes'] == pytest.approx(expected_signes + 2000.0)


def test_compute_bu_amounts_handles_empty_df():
    out = _compute_bu_amounts(pd.DataFrame(), 2026, 'TRAVAUX')
    assert out == {"signes": 0.0, "potentiels": 0.0, "envoyes": 0.0}


# ---------------------------------------------------------------------------
# _compute_maintenance_entries
# ---------------------------------------------------------------------------

def test_compute_maintenance_entries_filters_won_year_bu_maintenance(sample_df):
    entries = _compute_maintenance_entries(sample_df, 2026)
    assert len(entries) == 1
    e = entries[0]
    assert e["nom"].startswith("(E) - Maintenance Won 1")
    assert e["mois_signature"] == "Mars"
    assert e["mois_demarrage"] == "Mars"


def test_compute_maintenance_entries_uses_montant_total_year_for_prod_amount(sample_df):
    entries = _compute_maintenance_entries(sample_df, 2026)
    assert entries[0]["montant_ht_prod"] == pytest.approx(7088 / 12 * 10)
    assert entries[0]["montant_ht"] == pytest.approx(7088.10)


def test_compute_maintenance_entries_excludes_previous_year_signature(sample_df):
    # M-W-OLD has signature_date 2025-11-01 → must be excluded
    entries = _compute_maintenance_entries(sample_df, 2026)
    titles = [e["nom"] for e in entries]
    assert all("Last Year" not in t for t in titles)


def test_compute_maintenance_entries_handles_missing_columns():
    df = pd.DataFrame([{
        "id_devis": "X", "title": "Bare", "statut_clean": "gagné", "final_bu": "MAINTENANCE",
    }])
    entries = _compute_maintenance_entries(df, 2026)
    # No signature_date / date / date_effective_won → cannot resolve year → excluded
    assert entries == []


# ---------------------------------------------------------------------------
# build_budget_workbook (integration)
# ---------------------------------------------------------------------------

def test_build_budget_workbook_produces_two_sheets_with_expected_headers(sample_df):
    import openpyxl

    blob = build_budget_workbook(
        year=2026,
        df_processed=sample_df,
        portefeuille_debut_annee=996697.45,
        portefeuille_running=1043709.16,
        today=date(2026, 3, 16),
    )
    assert isinstance(blob, (bytes, bytearray))
    assert len(blob) > 0

    wb = openpyxl.load_workbook(BytesIO(blob), data_only=False)
    assert wb.sheetnames == ["Budget 2026 avec légende", "Maintenance"]

    ws1 = wb["Budget 2026 avec légende"]
    # Légende block
    assert ws1["D11"].value == "Légende"
    assert "Devis Signés" in (ws1["D12"].value or "")
    assert LEGEND_TEXT_TEMPLATE.format(year=2026).split("\n")[0] in ws1["D12"].value

    # Date stamp
    assert ws1["D16"].value == "Au 16/03/2026"

    # BU header band
    assert ws1["D17"].value == "Projection 2026"
    assert ws1["E17"].value == "CONCEPTION"
    assert ws1["H17"].value == "TRAVAUX"
    assert ws1["K17"].value == "MAINTENANCE"
    assert ws1["N17"].value == "TOTAL"

    # Subheaders
    assert ws1["E18"].value == "Devis Signés"
    assert ws1["F18"].value == "Devis Potentiels"
    assert ws1["G18"].value == "Devis Envoyés"
    assert ws1["K18"].value == "Nouveaux contrats 2026"

    # Numeric values come from sample_df aggregation
    assert ws1["E19"].value == pytest.approx(90.0)   # CONCEPTION signes
    assert ws1["F19"].value == pytest.approx(90.0)   # CONCEPTION potentiels
    assert ws1["G19"].value == pytest.approx(270.0)  # CONCEPTION envoyes

    # Formulas in row 20 / row 22 / row 25
    assert ws1["E20"].value == "=E19+F19"
    assert ws1["L22"].value == "=L19+L21"
    assert ws1["E25"].value == "=E23-E20"
    assert ws1["N25"].value == "=E25+H25+K25"

    # Portefeuille values
    assert ws1["L21"].value == pytest.approx(1043709.16)

    ws2 = wb["Maintenance"]
    assert ws2["B2"].value == "Entrées/Sortie Portefeuille sites"
    assert ws2["B3"].value == "Nom"
    assert ws2["C3"].value == "Montant HT Prod 2026"
    assert ws2["D3"].value == "Montant HT"
    assert ws2["E3"].value == "Mois signature"
    assert ws2["F3"].value == "Mois démarrage"
    assert ws2["A3"].value == "Au 16/03/26"
    assert ws2["C4"].value == pytest.approx(996697.45)
    # First (and only) entry row
    assert (ws2["B5"].value or "").startswith("(E) - Maintenance Won 1")
    assert ws2["E5"].value == "Mars"


def test_build_budget_workbook_handles_missing_portefeuille(sample_df):
    import openpyxl

    blob = build_budget_workbook(
        year=2027,  # no entretien début for 2027
        df_processed=sample_df,
        portefeuille_debut_annee=None,
        portefeuille_running=None,
        today=date(2027, 1, 5),
    )
    wb = openpyxl.load_workbook(BytesIO(blob), data_only=False)
    ws1 = wb["Budget 2027 avec légende"]
    # Empty portefeuille cell, but layout still intact
    assert ws1["L21"].value in (None, "")
    assert ws1["L22"].value == "=L19+L21"
